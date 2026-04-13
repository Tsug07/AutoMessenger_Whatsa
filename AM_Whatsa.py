import json
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import time
import random
import os
import psutil
import re
from collections import OrderedDict
import openpyxl
import customtkinter as ctk
from selenium import webdriver
from PIL import Image, ImageTk  # Add PIL for image handling
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
import requests
from dotenv import load_dotenv
load_dotenv()

"""
AutoMessenger WhatsApp - Ferramenta de automação para envio de mensagens via WhatsApp Web.
Supports multiple models with customizable Excel structures and messages.
"""

# Configuração do tema do customtkinter
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

# Variável global para o tema atual
tema_atual = "Dark"

def alternar_tema():
    """Alterna entre tema Dark e Light"""
    global tema_atual
    if tema_atual == "Dark":
        tema_atual = "Light"
        ctk.set_appearance_mode("Light")
    else:
        tema_atual = "Dark"
        ctk.set_appearance_mode("Dark")

    # Atualizar texto do botão se existir
    if 'botao_tema' in globals() and botao_tema is not None:
        icone = "☀" if tema_atual == "Dark" else "🌙"
        botao_tema.configure(text=icone)

# Variáveis globais
cancelar = False
log_file_path = None
anexo_habilitado = None  # Variável para checkbox de anexo
caminho_anexo = None  # Variável para caminho do arquivo anexo
agendamento_ativo = None  # Timer do agendamento
contagem_regressiva_ativa = False  # Flag para contagem regressiva
data_hora_agendada = None  # Data/hora do agendamento
perfil_selecionado = None  # Perfil do Chrome (1 ou 2)
driver_agendamento = None  # Driver do Chrome para agendamento
keep_alive_ativo = False  # Flag para keep-alive
KEEP_ALIVE_INTERVALO = 30 * 60 * 1000  # 30 minutos em milissegundos
INTERVALO_MIN = 7 * 60  # Mínimo 7 minutos entre cada envio (em segundos)
INTERVALO_MAX = 10 * 60  # Máximo 10 minutos entre cada envio (em segundos)

# Modelos suportados
MODELOS = {
    "ONE": {
        "colunas": ["Codigo", "Nome", "Numero", "Caminho"],
        "mensagem_padrao": "ONEmessage"
    },
    "ALL": {
        "colunas": ["Codigo", "Empresa", "Contato Onvio", "Grupo Onvio", "CNPJ", "Telefone"],
        "mensagem_padrao": "Mensagem Padrão"
    },
    "ALL_info": {
        "colunas": ["Codigo", "Nome", "Numero"],
        "colunas_opcionais": ["CNPJ", "Competencia", "Info_Extra"],
        "mensagem_padrao": "ALLinfo"
    },
    "Cobranca": {
        "colunas": ["Codigo", "Nome", "Numero", "Valor da Parcela", "Data de Vencimento", "Carta de Aviso"],
        "mensagem_padrao": "Cobranca"
    },
    "ComuniCertificado": {
       "colunas": ["Codigo", "Nome", "Numero", "CNPJ", "Vencimento", "Carta de Aviso"],
        "mensagem_padrao": "Cobranca"
    }
}

def esperar_carregamento_completo(driver):
    try:
        WebDriverWait(driver, 60).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        atualizar_log("Página completamente carregada.")
        return True
    except Exception as e:
        atualizar_log(f"Erro ao esperar carregamento: {str(e)}", cor="vermelho")
        return False

def formatar_telefone_whatsapp(telefone):
    """Formata o número de telefone para uso na URL do WhatsApp.
    Remove caracteres não numéricos. Se não começar com 55, adiciona o prefixo do Brasil."""
    telefone = re.sub(r'\D', '', str(telefone))
    if not telefone.startswith('55'):
        telefone = '55' + telefone
    return telefone

def aguardar_intervalo_envio():
    """Aguarda um intervalo aleatório entre envios com contagem regressiva no log."""
    intervalo = random.randint(INTERVALO_MIN, INTERVALO_MAX)
    minutos = intervalo // 60
    segundos_restantes = intervalo % 60
    if segundos_restantes > 0:
        atualizar_log(f"Aguardando {minutos}min {segundos_restantes}s até o próximo envio...")
    else:
        atualizar_log(f"Aguardando {minutos}min até o próximo envio...")
    for restante in range(intervalo, 0, -30):
        if cancelar:
            return
        m = restante // 60
        s = restante % 60
        atualizar_log(f"  {m}min {s}s restantes...")
        time.sleep(min(30, restante))


def verificar_numero_invalido(driver, telefone_formatado):
    """Verifica se apareceu o popup 'O número não está no WhatsApp' e clica em OK."""
    SELETOR_BTN_OK = "#app > div > div > span:nth-child(3) > div > span > div > div > div > div > div > div.x78zum5.x8hhl5t.x13a6bvl.x13crsa5.x1gabggj.x18d9i69.xaso8d8.xp4054r.xuxw1ft > div > button"
    try:
        botao_ok = WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, SELETOR_BTN_OK))
        )
        botao_ok.click()
        atualizar_log(f"Número {telefone_formatado} não está no WhatsApp! Pulando...", cor="vermelho")
        time.sleep(1)
        return True
    except Exception:
        return False


def navegar_para_contato_whatsapp(driver, telefone):
    """Navega para o chat do WhatsApp Web usando URL direta.
    Se já estiver no WhatsApp Web, usa a URL direta do send. Caso contrário, passa pelo wa.me."""
    try:
        telefone_formatado = formatar_telefone_whatsapp(telefone)
        url_atual = driver.current_url
        atualizar_log(f"Navegando para WhatsApp: {telefone_formatado}...")

        # Se já está no WhatsApp Web, navegar direto pela URL send
        if "web.whatsapp.com" in url_atual:
            url_direta = f"https://web.whatsapp.com/send?phone={telefone_formatado}"
            atualizar_log(f"Já no WhatsApp Web, navegando direto...")
            driver.get(url_direta)
            time.sleep(8)

            if not esperar_carregamento_completo(driver):
                return False

            # Verificar se o número não está no WhatsApp
            if verificar_numero_invalido(driver, telefone_formatado):
                return False

            atualizar_log(f"Chat aberto para {telefone_formatado}.", cor="azul")
            return True

        # Primeira vez: precisa passar pelo wa.me
        url = f"https://wa.me/{telefone_formatado}"
        atualizar_log(f"URL: {url}")

        # Guardar a aba original
        aba_original = driver.current_window_handle
        abas_antes = set(driver.window_handles)

        driver.get(url)
        time.sleep(5)

        if not esperar_carregamento_completo(driver):
            return False

        # Clicar no botão "Continuar para o WhatsApp Web"
        botao_continuar = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="action-button"][.//span[contains(text(), "WhatsApp Web")]]'))
        )
        botao_continuar.click()
        atualizar_log("Botão 'Continuar para o WhatsApp Web' clicado.")
        time.sleep(5)

        # Aguardar nova aba abrir e trocar para ela
        WebDriverWait(driver, 15).until(lambda d: len(d.window_handles) > len(abas_antes))
        abas_novas = set(driver.window_handles) - abas_antes
        nova_aba = abas_novas.pop()
        driver.switch_to.window(nova_aba)
        atualizar_log("Trocado para nova aba do WhatsApp Web.")

        # Fechar a aba antiga (página de confirmação)
        driver.switch_to.window(aba_original)
        driver.close()
        driver.switch_to.window(nova_aba)

        # Aguardar o chat carregar
        time.sleep(8)
        if not esperar_carregamento_completo(driver):
            return False

        # Verificar se o número não está no WhatsApp
        if verificar_numero_invalido(driver, telefone_formatado):
            return False

        atualizar_log(f"Chat aberto para {telefone_formatado}.", cor="azul")
        return True
    except Exception as e:
        atualizar_log(f"Erro ao navegar para contato WhatsApp ({telefone}): {str(e)}", cor="vermelho")
        return False

MENSAGEM_AVISO_NUMERO = (
    "Este número é utilizado apenas para envio de informações.\n"
    "\n"
    "❌ Não respondemos por aqui!\n"
    "\n"
    "📲 Para atendimento, entre em contato com nossa equipe pelo número oficial: (24) 99921-2350."
)


def digitar_e_enviar(driver, texto):
    """Localiza a caixa de mensagem, digita o texto e clica em enviar. Retorna True se sucesso."""
    caixa_msg = None
    for tentativa in range(3):
        try:
            caixa_msg = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '#main footer div.lexical-rich-text-input p'))
            )
            caixa_msg.click()
            break
        except Exception:
            atualizar_log(f"Tentativa {tentativa + 1}/3: caixa de mensagem não encontrada, aguardando...", cor="azul")
            time.sleep(5)
    if not caixa_msg:
        atualizar_log("Não foi possível encontrar a caixa de mensagem.", cor="vermelho")
        return False

    # Digitar texto caractere por caractere para simular digitação humana
    texto_limpo = texto.strip()
    driver.execute_script("arguments[0].focus();", caixa_msg)
    for char in texto_limpo:
        if char == '\n':
            # Shift+Enter para nova linha
            driver.execute_script(
                """
                var element = arguments[0];
                var br = new KeyboardEvent('keydown', {key: 'Enter', code: 'Enter', keyCode: 13, shiftKey: true, bubbles: true});
                element.dispatchEvent(br);
                """,
                caixa_msg
            )
        else:
            driver.execute_script("document.execCommand('insertText', false, arguments[0]);", char)
        time.sleep(random.uniform(0.02, 0.05))
    time.sleep(random.uniform(0.8, 2.0))

    botao_enviar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span/div/div/div/div[4]/div/span/button'))
    )
    time.sleep(random.uniform(0.3, 1.0))
    botao_enviar.click()
    time.sleep(random.uniform(2.0, 4.0))
    return True


def enviar_mensagem(driver, telefone, mensagem, codigo, identificador, modelo=None, caminhos=None, enviar_aviso=True):
    """Envia mensagem via WhatsApp Web navegando pela URL direta do contato."""
    try:
        if not navegar_para_contato_whatsapp(driver, telefone):
            atualizar_log(f"Falha ao abrir chat do WhatsApp para {telefone}.", cor="vermelho")
            return False

        tem_mensagem = mensagem and mensagem.strip()

        if tem_mensagem:
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return False

            # Enviar mensagem (já inclui o aviso de número no final)
            atualizar_log("Enviando mensagem...")
            if not digitar_e_enviar(driver, mensagem):
                return False
            atualizar_log("Mensagem enviada!", cor="azul")
        else:
            if not caminhos:
                atualizar_log("Erro: Sem mensagem e sem arquivos para enviar.", cor="vermelho")
                return False
            atualizar_log("Sem mensagem de texto, enviando apenas anexo...", cor="azul")

        # TODO: Implementar envio de anexos via seletores do WhatsApp Web
        if caminhos:
            atualizar_log("TODO: Envio de anexos ainda não implementado para WhatsApp Web.", cor="vermelho")

        atualizar_log(f"\nAviso enviado para {telefone}, {codigo} - {identificador}.\n", cor="verde")
        return True
    except Exception as e:
        atualizar_log(f"Erro ao enviar mensagem para {telefone}: {str(e)}", cor="vermelho")
        return False

# Funções de Navegação e Automação (reutilizadas do main.py e prorcontrato.py)
def obter_perfil_chrome():
    """Retorna o número do perfil baseado na seleção do usuário"""
    return perfil_selecionado.get() if perfil_selecionado else "1"

def obter_user_data_dir():
    """Retorna o diretório de dados do Chrome baseado no perfil selecionado."""
    perfil = obter_perfil_chrome()
    if perfil == "Teste":
        return os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data")
    return rf"C:\PerfisChrome\automacao_perfil{perfil}"

def abrir_chrome_com_url(url):
    perfil = obter_perfil_chrome()
    user_data_dir = obter_user_data_dir()

    if perfil == "Teste":
        return abrir_chrome_teste_com_url(url)

    # Encerra apenas o Chrome do perfil de automação
    encerrar_processos_chrome()

    # Criar diretório se não existir
    if not os.path.exists(user_data_dir):
        os.makedirs(user_data_dir, exist_ok=True)
        atualizar_log(f"Diretório do perfil {perfil} criado.", cor="azul")
        atualizar_log("Por favor, faça login na página aberta para continuar.", cor="azul")

    atualizar_log(f"Usando perfil: {perfil} ({user_data_dir})", cor="azul")

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-translate")
    chrome_options.add_argument("--lang=pt-BR")
    chrome_options.add_argument("--enable-javascript")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")

    service = Service(ChromeDriverManager().install())
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
        driver.set_page_load_timeout(180)
        driver.get(url)
        atualizar_log(f"Chrome aberto com a URL: {url}")

        return driver
    except Exception as e:
        atualizar_log(f"Erro ao abrir o Chrome: {str(e)}")
        return None

def abrir_chrome_teste_com_url(url):
    """Abre Chrome de teste copiando dados de sessão do perfil padrão para um perfil limpo."""
    import shutil

    # Fechar TUDO relacionado ao Chrome (comentado para evitar fechar Chrome pessoal)
    # atualizar_log("Fechando todos os processos Chrome...", cor="azul")
    # os.system("taskkill /f /im chrome.exe >nul 2>&1")
    # os.system("taskkill /f /im chromedriver.exe >nul 2>&1")
    # for proc in psutil.process_iter(['name']):
    #     try:
    #         if proc.info['name'] and 'chrome' in proc.info['name'].lower():
    #             proc.kill()
    #     except (psutil.NoSuchProcess, psutil.AccessDenied):
    #         pass
    # time.sleep(5)

    # Copiar dados de sessão do perfil padrão para o perfil de teste
    perfil_padrao = os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data\Profile 1")
    perfil_teste = r"C:\PerfisChrome\automacao_teste\Default"
    os.makedirs(perfil_teste, exist_ok=True)

    # Copiar arquivos essenciais de sessão (cookies, localStorage, IndexedDB do WhatsApp)
    arquivos_sessao = [
        "Cookies", "Cookies-journal",
        "Login Data", "Login Data-journal",
        "Web Data", "Web Data-journal",
        "Preferences", "Secure Preferences",
    ]
    pastas_sessao = ["Local Storage", "IndexedDB", "Service Worker"]

    for arquivo in arquivos_sessao:
        src = os.path.join(perfil_padrao, arquivo)
        dst = os.path.join(perfil_teste, arquivo)
        if os.path.exists(src):
            try:
                shutil.copy2(src, dst)
            except Exception:
                pass

    for pasta in pastas_sessao:
        src = os.path.join(perfil_padrao, pasta)
        dst = os.path.join(perfil_teste, pasta)
        if os.path.exists(src):
            try:
                if os.path.exists(dst):
                    shutil.rmtree(dst)
                shutil.copytree(src, dst)
            except Exception:
                pass

    atualizar_log("Dados de sessão copiados do perfil padrão.", cor="azul")

    # Abrir Chrome com o perfil de teste (limpo, sem conflito)
    user_data_dir = r"C:\PerfisChrome\automacao_teste"
    atualizar_log(f"Usando perfil: Teste ({user_data_dir})", cor="azul")

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-translate")
    chrome_options.add_argument("--lang=pt-BR")
    chrome_options.add_argument("--enable-javascript")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")

    service = Service(ChromeDriverManager().install())
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
        driver.set_page_load_timeout(180)
        driver.get(url)
        atualizar_log(f"Chrome aberto com perfil de teste. URL: {url}", cor="verde")
        return driver
    except Exception as e:
        atualizar_log(f"Erro ao abrir Chrome: {str(e)}", cor="vermelho")
        return None

def encerrar_processos_chrome():
    """Encerra apenas os processos Chrome do perfil selecionado"""
    perfil = obter_perfil_chrome()
    encerrou_algum = False
    for proc in psutil.process_iter(['name', 'cmdline']):
        if proc.info['name'] == 'chrome.exe':
            try:
                cmdline = proc.info['cmdline'] or []
                cmdline_str = ' '.join(cmdline)
                # Encerra apenas o Chrome do diretório do perfil selecionado
                # Verifica com barras normais e invertidas
                if f'automacao_perfil{perfil}' in cmdline_str:
                    proc.terminate()
                    atualizar_log(f"Processo Chrome (Perfil {perfil}) encerrado (PID: {proc.pid}).")
                    encerrou_algum = True
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
    if encerrou_algum:
        time.sleep(2)

# Funções de Dados
def validar_excel(caminho, modelo):
    try:
        wb = openpyxl.load_workbook(caminho)
        sheet = wb.active
        colunas_excel = [cell.value for cell in sheet[1]]
        colunas_esperadas = MODELOS[modelo]["colunas"]

        # Para ALL_info, aceitar colunas opcionais adicionais
        if modelo == "ALL_info":
            colunas_opcionais = MODELOS[modelo].get("colunas_opcionais", [])
            # Verifica se as colunas obrigatórias estão presentes
            colunas_obrigatorias = colunas_esperadas[:3]  # Codigo, Empresa, Telefone
            if colunas_excel[:3] != colunas_obrigatorias:
                messagebox.showerror("Erro", f"O Excel não corresponde ao modelo {modelo}. Colunas obrigatórias: {colunas_obrigatorias}")
                return False
            # Verifica se as colunas extras são válidas (opcionais)
            colunas_extras = colunas_excel[3:]
            for col in colunas_extras:
                if col and col not in colunas_opcionais:
                    messagebox.showwarning("Aviso", f"Coluna '{col}' não reconhecida. Colunas opcionais válidas: {colunas_opcionais}")
            atualizar_log(f"Colunas detectadas: {colunas_excel}")
            return True

        # Aceitar Excel com colunas extras (ex: CNPJ) desde que as colunas esperadas estejam presentes no início
        if colunas_excel[:len(colunas_esperadas)] != colunas_esperadas:
            messagebox.showerror("Erro", f"O Excel não corresponde ao modelo {modelo}. Esperado: {colunas_esperadas}")
            return False
        if len(colunas_excel) > len(colunas_esperadas):
            atualizar_log(f"Colunas extras ignoradas: {colunas_excel[len(colunas_esperadas):]}")
        return True
    except Exception as e:
        atualizar_log(f"Erro ao validar Excel: {str(e)}", cor="vermelho")
        return False

def ler_dados_excel(caminho_excel, modelo, linha_inicial=2):
    try:
        wb = openpyxl.load_workbook(caminho_excel)
        sheet = wb.active
        dados = {}
        colunas = MODELOS[modelo]["colunas"]

        # Para ALL_info, detectar colunas dinamicamente
        colunas_excel = []
        if modelo == "ALL_info":
            colunas_excel = [cell.value for cell in sheet[1]]
            dados['_colunas_detectadas'] = colunas_excel  # Armazenar para uso posterior

        for row in sheet.iter_rows(min_row=linha_inicial, values_only=True):
            if row and len(row) >= len(colunas):
                codigo = row[0]

                if modelo == "Cobranca":
                    # Colunas: Código, Empresa, Telefone, Valor da Parcela, Data de Vencimento, Carta de Aviso
                    codigo, nome, telefone, valores, vencimentos, cartas = row[:6]

                    if not isinstance(cartas, (int, float)) or not 1 <= int(cartas) <= 7:
                        atualizar_log(f"Linha ignorada: Carta de aviso inválida ({cartas}) na linha {row[0]}", cor="vermelho")
                        continue
                    if codigo in dados:
                        dados[codigo]['detalhes'].append({
                            'valores': valores,
                            'vencimentos': vencimentos
                        })
                    else:
                        dados[codigo] = {
                            'nome': nome,
                            'telefone': str(telefone),
                            'detalhes': [{
                                'valores': valores,
                                'vencimentos': vencimentos
                            }],
                            'cartas': cartas
                        }

                elif modelo == "ComuniCertificado":
                    # Colunas: Codigo, Empresa, Telefone, CNPJ, Vencimento, Carta de Aviso
                    codigo, nome, telefone, cnpj, vencimentos, cartas = row[:6]
                    dados[codigo] = {
                        'nome': nome,
                        'telefone': str(telefone),
                        'cnpj': cnpj,
                        'vencimentos': vencimentos,
                        'cartas': cartas
                    }

                elif modelo == "ONE":
                    # Colunas: Código, Empresa, Telefone, Caminho
                    empresa, telefone, caminho = row[1:4]
                    telefone = str(telefone) if telefone is not None else ""
                    # Agrupar por telefone
                    if telefone in dados:
                        dados[telefone]['empresas'].append({
                            'codigo': codigo,
                            'empresa': empresa,
                            'caminho': caminho
                        })
                    else:
                        dados[telefone] = {
                            'telefone': telefone,
                            'empresas': [{
                                'codigo': codigo,
                                'empresa': empresa,
                                'caminho': caminho
                            }]
                        }

                elif modelo == "ALL_info":
                    # Colunas obrigatórias: Codigo, Empresa, Telefone
                    empresa, telefone = row[1:3]
                    telefone = str(telefone) if telefone is not None else ""

                    empresa_data = {
                        'codigo': codigo,
                        'empresa': empresa
                    }
                    info_extra = {}

                    # Mapear colunas extras baseado no header (a partir do índice 3)
                    if colunas_excel:
                        for idx, col_name in enumerate(colunas_excel[3:], start=3):
                            if col_name and idx < len(row):
                                valor = row[idx]
                                col_name_upper = str(col_name).strip().upper()
                                if col_name_upper == "COMPETENCIA":
                                    info_extra['competencia'] = str(valor) if valor is not None else ""
                                elif col_name_upper == "CNPJ":
                                    empresa_data['cnpj'] = str(valor) if valor is not None else ""
                                elif col_name_upper == "INFO_EXTRA":
                                    empresa_data['info_extra'] = str(valor) if valor is not None else ""

                    # Agrupar por telefone
                    if telefone in dados and telefone != '_colunas_detectadas':
                        dados[telefone]['empresas'].append(empresa_data)
                        for key, val in info_extra.items():
                            if key not in dados[telefone]:
                                dados[telefone][key] = val
                    else:
                        dados[telefone] = {
                            'telefone': telefone,
                            'empresas': [empresa_data],
                            **info_extra
                        }

                else:  # Modelo ALL
                    # Colunas: Codigo, Empresa, Contato Onvio, Grupo Onvio, CNPJ, Telefone
                    empresa = row[1]
                    telefone = str(row[5]) if row[5] is not None else ""
                    # Agrupar por telefone
                    if telefone in dados:
                        dados[telefone]['empresas'].append({
                            'codigo': codigo,
                            'empresa': empresa
                        })
                    else:
                        dados[telefone] = {
                            'telefone': telefone,
                            'empresas': [{
                                'codigo': codigo,
                                'empresa': empresa
                            }]
                        }
            else:
                atualizar_log(f"Linha ignorada: {row}")
        return dados if dados else None
    except Exception as e:
        atualizar_log(f"Erro ao ler Excel: {str(e)}", cor="vermelho")
        return None

def extrair_dados(dados, modelo):
    codigos, telefones = [], []

    if modelo == "Cobranca":
        nome, valores, vencimentos, cartas = [], [], [], []
        for cod, info in dados.items():
            codigos.append(cod)
            nome.append(info['nome'])
            telefones.append(info['telefone'])

            valor_total = []
            vencimento_total = []
            for detalhe in info['detalhes']:
                valor_total.append(detalhe['valores'])
                vencimento_total.append(detalhe['vencimentos'])

            valores.append(valor_total)
            vencimentos.append(vencimento_total)
            cartas.append(info['cartas'])

        return codigos, nome, telefones, valores, vencimentos, cartas

    elif modelo == "ComuniCertificado":
        nome, cnpjs, vencimentos, cartas = [], [], [], []
        for cod, info in dados.items():
            codigos.append(cod)
            nome.append(info['nome'])
            telefones.append(info['telefone'])
            cnpjs.append(info['cnpj'])
            vencimentos.append(info['vencimentos'])
            cartas.append(info['cartas'])

        return codigos, nome, telefones, cnpjs, vencimentos, cartas

    elif modelo == "ONE":
        telefones_lista, empresas_lista, caminhos_lista = [], [], []
        for chave, info in dados.items():
            telefones_lista.append(info['telefone'])
            empresas = [(emp['codigo'], emp['empresa'], emp['caminho']) for emp in info['empresas']]
            empresas_lista.append(empresas)
            caminhos_lista.append([emp['caminho'] for emp in info['empresas']])
        return telefones_lista, empresas_lista, caminhos_lista

    elif modelo == "ALL_info":
        telefones_lista, empresas_lista, extras = [], [], []
        for chave, info in dados.items():
            if chave == '_colunas_detectadas':
                continue
            telefones_lista.append(info['telefone'])
            extra_info = {
                'competencia': info.get('competencia', ''),
            }
            extras.append(extra_info)
            empresas = []
            for emp in info['empresas']:
                emp_data = {
                    'codigo': emp['codigo'],
                    'empresa': emp['empresa'],
                    'cnpj': emp.get('cnpj', ''),
                    'info_extra': emp.get('info_extra', '')
                }
                empresas.append(emp_data)
            empresas_lista.append(empresas)
        return telefones_lista, empresas_lista, extras

    else:  # Modelo ALL
        telefones_lista, empresas_lista = [], []
        for chave, info in dados.items():
            telefones_lista.append(info['telefone'])
            empresas = [(emp['codigo'], emp['empresa']) for emp in info['empresas']]
            empresas_lista.append(empresas)
        return telefones_lista, empresas_lista
    
def formatar_cnpj(cnpj):
    # Remover caracteres não numéricos
    cnpj = ''.join(filter(str.isdigit, cnpj))
    
    # Verificar se o CNPJ tem 14 dígitos
    if len(cnpj) != 14:
        raise ValueError("CNPJ deve conter 14 dígitos")
    
    # Formatar o CNPJ no padrão: XX.XXX.XXX/XXXX-XX
    cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

    return cnpj_formatado

# Funções de Mensagem
def carregar_mensagens():
    try:
        with open("mensagens.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {
            "Mensagem Padrão": "Teste Desconsiderando mensagem",
            "Prorrogação Contrato": "Prezado cliente,\nEspero que estejam bem.\n\nGostaríamos de informar que o contrato de experiência das seguintes pessoas está preste a vencer:\n\n{pessoas_vencimentos}\n\nPara darmos prosseguimento aos devidos registros, solicitamos a gentileza de nos confirmar se haverá prorrogação do contrato ou se ele será encerrado nesta data.\n\nCaso não recebamos um retorno, entenderemos que a prorrogação será realizada automaticamente.\n\nAgradecemos sua atenção.\n\nAtenciosamente,\n\nEquipe DP - C&S."
        }

def salvar_mensagens(mensagens):
    with open("mensagens.json", "w", encoding="utf-8") as f:
        json.dump(mensagens, f, ensure_ascii=False, indent=4)

def mensagem_padrao(modelo, pessoas=None, vencimentos=None, valores=None, carta=None, cnpj=None, nome_empresa=None, competencia=None, empresas_info=None):
    mensagens = carregar_mensagens()
    msg = mensagens.get(mensagem_selecionada.get(), MODELOS[modelo]["mensagem_padrao"])
    
    # if modelo == "ProrContrato" and pessoas and vencimentos:
    #     pv = "\n".join([f"{p} se encerrará em {v}" for p, v in zip(pessoas, vencimentos)])
    #     msg = msg.format(pessoas_vencimentos=pv)
    if modelo == "Cobranca" and valores and vencimentos and nome_empresa and carta is not None:
        # Formatar valores com vírgula como separador decimal
        valores_formatados = [f"{valor:.2f}".replace('.', ',') for valor in valores]
        total_formatado = f"{sum(valores):.2f}".replace('.', ',')
        # Formatar parcelas
        parcelas = "\n".join([f"Valor: R$ {valor} | Vencimento: {venc}" for valor, venc in zip(valores_formatados, vencimentos)])
        # Selecionar a mensagem com base no número da carta
        msg_key = f"Cobranca_{carta}" if f"Cobranca_{carta}" in mensagens else "Cobranca_1"  # Fallback para carta 1
        msg = mensagens.get(msg_key, mensagens.get("Cobranca_1", "Mensagem de cobrança padrão não encontrada."))
        msg = msg.format(nome=nome_empresa, parcelas=parcelas, total=total_formatado)
    
    elif modelo == "ComuniCertificado":
        cnpj_formatado = formatar_cnpj(cnpj)
         # Selecionar a mensagem com base no número da carta
        msg_key = f"Certificado_{carta}" if f"Certificado_{carta}" in mensagens else "Certificado_1"  # Fallback para carta 1
        msg = mensagens.get(msg_key, mensagens.get("Certificado_1", "Mensagem de cobrança padrão não encontrada."))
        msg = msg.format(nome=nome_empresa, cnpj_formatado=cnpj_formatado, datas=vencimentos)
    
    elif modelo in ["ONE", "ALL", "ALL_info"]:
        # normaliza nome_empresa para lista de nomes
        if isinstance(nome_empresa, list):
            nomes_empresas = nome_empresa
        elif nome_empresa is None:
            nomes_empresas = []
        else:
            nomes_empresas = [nome_empresa]

        # Pegar a mensagem selecionada pelo usuário
        msg_selecionada = mensagem_selecionada.get()

        # Verificar se é uma mensagem que NÃO usa dados (sem placeholders)
        if msg_selecionada == "ONEmessage":
            # Mensagem simples sem dados dinâmicos
            msg = mensagens.get(msg_selecionada, "Mensagem padrão não encontrada.")
        else:
            # Mensagem com dados (Parabens_Regularizado, ALLinfo, SemReceita, etc.)
            if len(nomes_empresas) > 1:
                # Múltiplas empresas - usa versão _multi
                msg_key = f"{msg_selecionada}_multi" if f"{msg_selecionada}_multi" in mensagens else msg_selecionada
                msg = mensagens.get(msg_key, mensagens.get(msg_selecionada, "Mensagem padrão não encontrada."))

                # Verificar se a mensagem precisa de empresas com CNPJ
                if empresas_info and '{empresas_cnpj}' in msg:
                    # Formatar lista de empresas com CNPJ
                    lista_empresas_cnpj = []
                    for emp in empresas_info:
                        cnpj_emp = emp.get('cnpj', '')
                        if cnpj_emp:
                            try:
                                cnpj_formatado = formatar_cnpj(cnpj_emp)
                            except ValueError:
                                cnpj_formatado = cnpj_emp
                            lista_empresas_cnpj.append(f". {emp['empresa']}, CNPJ {cnpj_formatado}")
                        else:
                            lista_empresas_cnpj.append(f". {emp['empresa']}")
                    empresas_cnpj_str = "\n".join(lista_empresas_cnpj)
                    try:
                        msg = msg.format(empresas_cnpj=empresas_cnpj_str, competencia=competencia if competencia else "")
                    except KeyError:
                        pass
                else:
                    # Formato padrão sem CNPJ
                    lista_empresas = "\n".join([f". {emp}" for emp in nomes_empresas])
                    # Tentar formatar com lista_empresas e competência, se falhar, enviar sem formatação
                    try:
                        if competencia:
                            msg = msg.format(empresas=lista_empresas, competencia=competencia)
                        else:
                            msg = msg.format(empresas=lista_empresas)
                    except KeyError:
                        pass
            else:
                # Uma única empresa
                msg = mensagens.get(msg_selecionada, "Mensagem padrão não encontrada.")
                nome_unico = nomes_empresas[0] if nomes_empresas else ""

                # Verificar se a mensagem precisa de CNPJ
                if empresas_info and '{cnpj}' in msg:
                    cnpj_emp = empresas_info[0].get('cnpj', '') if empresas_info else ''
                    if cnpj_emp:
                        try:
                            cnpj_formatado = formatar_cnpj(cnpj_emp)
                        except ValueError:
                            cnpj_formatado = cnpj_emp
                    else:
                        cnpj_formatado = ''
                    try:
                        msg = msg.format(nome=nome_unico, cnpj=cnpj_formatado, competencia=competencia if competencia else "")
                    except KeyError:
                        pass
                else:
                    # Tentar formatar com nome e competência, se falhar, enviar sem formatação
                    try:
                        if competencia:
                            msg = msg.format(nome=nome_unico, competencia=competencia)
                        else:
                            msg = msg.format(nome=nome_unico)
                    except KeyError:
                        pass
    # Anexar aviso de número ao final da mensagem
    msg += "\n\n" + MENSAGEM_AVISO_NUMERO
    return msg

# Funções de Interface
def selecionar_excel():
    arquivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if arquivo:
        caminho_excel.set(arquivo)
        modelo = modelo_selecionado.get()
        if modelo and not validar_excel(arquivo, modelo):
            caminho_excel.set("")
        else:
            atualizar_log(f"Arquivo Excel selecionado: {arquivo}")

def atualizar_mensagem_padrao(*args):
    modelo = modelo_selecionado.get()
    if modelo:
        mensagem_padrao_key = MODELOS[modelo]["mensagem_padrao"]
        if modelo == "Cobranca":
            mensagem_padrao_key = "Cobranca"
        elif modelo == "ComuniCertificado":
            mensagem_padrao_key = "Certificado"
        mensagem_selecionada.set(mensagem_padrao_key)

def iniciar_processamento():
    global cancelar, driver_agendamento, keep_alive_ativo
    cancelar = False

    # Verificar se há agendamento ativo - não permitir iniciar manualmente
    if agendamento_ativo or keep_alive_ativo:
        messagebox.showwarning("Atenção", "Há um agendamento ativo. Cancele o agendamento antes de iniciar manualmente.")
        return

    excel = caminho_excel.get()
    modelo = modelo_selecionado.get()
    if not excel or not modelo:
        messagebox.showwarning("Atenção", "Selecione um modelo e um arquivo Excel.")
        return
    try:
        linha = int(entrada_linha_inicial.get())
        if linha < 2:
            raise ValueError("Linha inicial deve ser >= 2")
    except ValueError:
        messagebox.showwarning("Atenção", "Linha inicial deve ser um número inteiro >= 2.")
        return
    atualizar_log("Iniciando processamento...", cor="azul")
    botao_iniciar.configure(state="disabled")
    botao_iniciar_chrome.configure(state="disabled")  # Desativar o botão de Chrome
    botao_agendar.configure(state="disabled")  # Desativar agendamento durante processamento
    inicializar_arquivo_log(modelo)
    thread = threading.Thread(target=processar_dados, args=(excel, modelo, linha))
    thread.start()

def formatar_tempo(tempo_inicio):
    """Calcula e formata o tempo decorrido desde tempo_inicio."""
    tempo_total = time.time() - tempo_inicio
    horas = int(tempo_total // 3600)
    minutos = int((tempo_total % 3600) // 60)
    segundos = int(tempo_total % 60)
    if horas > 0:
        return f"{horas}h {minutos}min {segundos}s"
    elif minutos > 0:
        return f"{minutos}min {segundos}s"
    else:
        return f"{segundos}s"

def estimar_tempo(num_envios, num_mensagens_extra=0):
    """Estima o tempo total de processamento baseado no número de envios (intervalos).
    num_envios: quantas vezes aguardar_intervalo_envio() será chamado
    num_mensagens_extra: mensagens extras dentro do mesmo chat (ex: múltiplas empresas no mesmo tel)
    """
    media_intervalo = (INTERVALO_MIN + INTERVALO_MAX) / 2
    media_envio = 5.0  # tempo médio por mensagem (navegação + digitação + envio)
    media_extra = 3.5  # tempo médio por mensagem extra no mesmo chat

    total_min = (num_envios * (INTERVALO_MIN + media_envio) + num_mensagens_extra * media_extra) / 60
    total_max = (num_envios * (INTERVALO_MAX + media_envio) + num_mensagens_extra * media_extra) / 60
    total_media = (total_min + total_max) / 2

    def fmt(minutos):
        h = int(minutos // 60)
        m = int(minutos % 60)
        if h > 0:
            return f"{h}h {m}min"
        return f"{m}min"

    atualizar_log(f"=" * 50, cor="azul")
    atualizar_log(f"ESTIMATIVA DE TEMPO", cor="azul")
    atualizar_log(f"Contatos/envios: {num_envios}" + (f" (+{num_mensagens_extra} msgs extras)" if num_mensagens_extra else ""), cor="azul")
    atualizar_log(f"Mínimo: ~{fmt(total_min)} | Média: ~{fmt(total_media)} | Máximo: ~{fmt(total_max)}", cor="azul")
    atualizar_log(f"=" * 50, cor="azul")


def processar_dados(excel, modelo, linha_inicial):
    # Iniciar timer de processamento
    tempo_inicio = time.time()
    atualizar_log("Timer iniciado.", cor="azul")

    url = "https://web.whatsapp.com"
    driver = abrir_chrome_com_url(url)
    if not driver:
        atualizar_log("Não foi possível abrir o Chrome. Processamento abortado.", cor="vermelho")
        finalizar_programa()
        return

    atualizar_log("Aguardando login no WhatsApp Web (escaneie o QR Code se necessário)...", cor="azul")
    time.sleep(15)
    dados = ler_dados_excel(excel, modelo, linha_inicial)
    if not dados:
        atualizar_log("Nenhum dado para processar.", cor="vermelho")
        return
    total_linhas = openpyxl.load_workbook(excel).active.max_row - linha_inicial + 1

    if modelo == "Cobranca":
        codigos, nomes, telefones, valores, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        # Agrupar índices por telefone (preservando ordem de primeira aparição)
        grupos_tel = OrderedDict()
        for i, tel in enumerate(telefones):
            tel_fmt = formatar_telefone_whatsapp(tel)
            if tel_fmt not in grupos_tel:
                grupos_tel[tel_fmt] = []
            grupos_tel[tel_fmt].append(i)

        num_grupos = len(grupos_tel)
        estimar_tempo(num_grupos, total_contatos - num_grupos)

        idx_processado = 0
        cobranca_enviados = 0
        for tel_fmt, indices in grupos_tel.items():
            if cancelar:
                atualizar_log(f"Processamento cancelado! Tempo decorrido: {formatar_tempo(tempo_inicio)}", cor="azul")
                return

            tel_original = telefones[indices[0]]
            atualizar_log(f"\nProcessando tel {tel_fmt}: {len(indices)} empresa(s)\n", cor="azul")

            # Navegar para o contato uma única vez
            if not navegar_para_contato_whatsapp(driver, tel_original):
                atualizar_log(f"Falha ao abrir chat para {tel_fmt}.", cor="vermelho")
                idx_processado += len(indices)
                continue

            # Enviar mensagem de cada empresa em sequência no mesmo chat
            for idx in indices:
                if cancelar:
                    atualizar_log(f"Processamento cancelado! Tempo decorrido: {formatar_tempo(tempo_inicio)}", cor="azul")
                    return
                cod, nome_emp, tel, p, v, carta = codigos[idx], nomes[idx], telefones[idx], valores[idx], vencimentos[idx], cartas[idx]
                idx_processado += 1
                linha_atual = linha_inicial + idx
                porcentagem = (idx_processado / total_contatos) * 100
                atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
                atualizar_log(f"Linha: {linha_atual}")
                atualizar_log(f"Empresa {cod} - {nome_emp}: Aviso nº: {carta}", cor="azul")
                mensagem = mensagem_padrao(modelo, valores=p, vencimentos=v, carta=carta, nome_empresa=nome_emp)
                atualizar_log("Enviando mensagem...")
                if digitar_e_enviar(driver, mensagem):
                    atualizar_log(f"Mensagem enviada para {nome_emp}!", cor="azul")
                    cobranca_enviados += 1
                    with open(log_file_path, 'a', encoding='utf-8') as f:
                        f.write(f"[{datetime.now()}] Mensagem enviada para {tel} - {cod} {nome_emp}\n")
                else:
                    atualizar_log(f"Falha ao enviar mensagem para {nome_emp}.", cor="vermelho")
                time.sleep(3)

            aguardar_intervalo_envio()

        notificar_discord_cobranca(cobranca_enviados)

    elif modelo == "ComuniCertificado":
        codigos, nomes, telefones, cnpjs, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        # Agrupar índices por telefone (preservando ordem de primeira aparição)
        grupos_tel = OrderedDict()
        for i, tel in enumerate(telefones):
            tel_fmt = formatar_telefone_whatsapp(tel)
            if tel_fmt not in grupos_tel:
                grupos_tel[tel_fmt] = []
            grupos_tel[tel_fmt].append(i)

        num_grupos = len(grupos_tel)
        estimar_tempo(num_grupos, total_contatos - num_grupos)

        idx_processado = 0
        certificado_enviados = 0
        for tel_fmt, indices in grupos_tel.items():
            if cancelar:
                atualizar_log(f"Processamento cancelado! Tempo decorrido: {formatar_tempo(tempo_inicio)}", cor="azul")
                return

            tel_original = telefones[indices[0]]
            atualizar_log(f"\nProcessando tel {tel_fmt}: {len(indices)} empresa(s)\n", cor="azul")

            # Navegar para o contato uma única vez
            if not navegar_para_contato_whatsapp(driver, tel_original):
                atualizar_log(f"Falha ao abrir chat para {tel_fmt}.", cor="vermelho")
                idx_processado += len(indices)
                continue

            # Enviar mensagem de cada empresa em sequência no mesmo chat
            for idx in indices:
                if cancelar:
                    atualizar_log(f"Processamento cancelado! Tempo decorrido: {formatar_tempo(tempo_inicio)}", cor="azul")
                    return
                cod, nome_emp, tel, c, v, carta = codigos[idx], nomes[idx], telefones[idx], cnpjs[idx], vencimentos[idx], cartas[idx]
                idx_processado += 1
                linha_atual = linha_inicial + idx
                porcentagem = (idx_processado / total_contatos) * 100
                atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
                atualizar_log(f"Linha: {linha_atual}")
                atualizar_log(f"Empresa {cod} - {nome_emp}: Aviso nº: {carta}", cor="azul")
                mensagem = mensagem_padrao(modelo, vencimentos=v, carta=carta, cnpj=c, nome_empresa=nome_emp)
                atualizar_log("Enviando mensagem...")
                if digitar_e_enviar(driver, mensagem):
                    atualizar_log(f"Mensagem enviada para {nome_emp}!", cor="azul")
                    certificado_enviados += 1
                    with open(log_file_path, 'a', encoding='utf-8') as f:
                        f.write(f"[{datetime.now()}] Mensagem enviada para {tel} - {cod} {nome_emp}\n")
                else:
                    atualizar_log(f"Falha ao enviar mensagem para {nome_emp}.", cor="vermelho")
                time.sleep(3)

            aguardar_intervalo_envio()

        notificar_discord_certificado(certificado_enviados)

    elif modelo == "ONE":
        telefones_lista, empresas_lista, caminhos_lista = extrair_dados(dados, modelo)
        total_contatos = len(telefones_lista)
        estimar_tempo(total_contatos)
        linha_atual = linha_inicial
        for i, (tel, empresas, caminhos) in enumerate(zip(telefones_lista, empresas_lista, caminhos_lista)):
            if cancelar:
                atualizar_log(f"Processamento cancelado! Tempo decorrido: {formatar_tempo(tempo_inicio)}", cor="azul")
                return
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando tel {tel}: {num_empresas} empresas\n", cor="azul")
            for cod, emp, _ in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")
            nomes_empresas = [emp for _, emp, _ in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas)
            identificador = ", ".join(nomes_empresas)
            if enviar_mensagem(driver, tel, mensagem, tel, identificador, modelo, caminhos):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] Mensagem enviada para {tel} com {num_empresas} arquivos\n")
            aguardar_intervalo_envio()
            linha_atual += num_empresas

    elif modelo == "ALL_info":
        telefones_lista, empresas_lista, extras = extrair_dados(dados, modelo)
        total_contatos = len(telefones_lista)
        estimar_tempo(total_contatos)
        linha_atual = linha_inicial
        for i, (tel, empresas, extra_info) in enumerate(zip(telefones_lista, empresas_lista, extras)):
            if cancelar:
                atualizar_log(f"Processamento cancelado! Tempo decorrido: {formatar_tempo(tempo_inicio)}", cor="azul")
                return
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            competencia = extra_info.get('competencia', '')
            log_extra = f" - Competencia: {competencia}" if competencia else ""
            atualizar_log(f"\nProcessando tel {tel}: {num_empresas} empresas{log_extra}\n", cor="azul")
            for emp in empresas:
                cnpj_log = f" - CNPJ: {emp.get('cnpj', '')}" if emp.get('cnpj') else ""
                atualizar_log(f"Empresa: {emp['codigo']} - {emp['empresa']}{cnpj_log}")
            nomes_empresas = [emp['empresa'] for emp in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas, competencia=competencia, empresas_info=empresas)
            identificador = ", ".join(nomes_empresas)
            if enviar_mensagem(driver, tel, mensagem, tel, identificador, modelo):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] Mensagem enviada para {tel} com {num_empresas} empresa(s){log_extra}\n")
            aguardar_intervalo_envio()
            linha_atual += num_empresas

    else:  # Modelo ALL
        telefones_lista, empresas_lista = extrair_dados(dados, modelo)
        total_contatos = len(telefones_lista)
        estimar_tempo(total_contatos)
        linha_atual = linha_inicial

        arquivo_anexo = None
        if anexo_habilitado and anexo_habilitado.get() and caminho_anexo and caminho_anexo.get():
            arquivo_anexo = caminho_anexo.get()
            if os.path.exists(arquivo_anexo):
                atualizar_log(f"Anexo configurado: {arquivo_anexo}", cor="azul")
            else:
                atualizar_log(f"Arquivo anexo não encontrado: {arquivo_anexo}", cor="vermelho")
                arquivo_anexo = None

        for i, (tel, empresas) in enumerate(zip(telefones_lista, empresas_lista)):
            if cancelar:
                atualizar_log(f"Processamento cancelado! Tempo decorrido: {formatar_tempo(tempo_inicio)}", cor="azul")
                return
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando tel {tel}: {num_empresas} empresas\n", cor="azul")
            for cod, emp in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")
            nomes_empresas = [emp for _, emp in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas)
            identificador = ", ".join(nomes_empresas)
            caminhos_envio = [arquivo_anexo] if arquivo_anexo else None
            if enviar_mensagem(driver, tel, mensagem, tel, identificador, modelo, caminhos_envio):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    anexo_info = " + anexo" if arquivo_anexo else ""
                    f.write(f"[{datetime.now()}] Mensagem enviada para {tel} com {num_empresas} empresa(s){anexo_info}\n")
            aguardar_intervalo_envio()
            linha_atual += num_empresas

    atualizar_progresso(100, "Concluído")
    atualizar_log(f"Tempo total de processamento: {formatar_tempo(tempo_inicio)}", cor="verde")
    atualizar_log("Processamento finalizado!", cor="verde")
    finalizar_programa()


def cancelar_processamento():
    global cancelar
    cancelar = True
    atualizar_log("Cancelando processamento...", cor="azul")
    botao_fechar.configure(state="normal")

def fechar_programa():
    global agendamento_ativo, keep_alive_ativo

    # Cancelar agendamento se estiver ativo
    if agendamento_ativo:
        agendamento_ativo.cancel()
        agendamento_ativo = None

    # Parar keep-alive e fechar Chrome do agendamento
    parar_keep_alive()
    fechar_chrome_agendamento()

    janela.quit()

def notificar_discord_cobranca(total_enviados):
    """Envia notificação ao Discord via webhook específico de cobrança."""
    webhook_url = os.getenv("DISCORD_WEBHOOK_COBRANCA")
    if not webhook_url:
        atualizar_log("Webhook de cobrança não configurado no .env.", cor="vermelho")
        return
    cargo = "<@&1299045096146079795>"
    payload = {
        "content": f"{cargo} **AutoMessenger WhatsApp - Cobrança** - Mensagens de cobrança enviadas! Total: {total_enviados} mensagem(ns) enviada(s)."
    }
    try:
        response = requests.post(webhook_url, json=payload)
        if response.status_code == 204:
            atualizar_log("Notificação de cobrança enviada ao Discord.", cor="verde")
        elif response.status_code == 403:
            atualizar_log("Falha ao notificar Discord cobrança: Webhook inválido ou deletado (403 Forbidden).", cor="vermelho")
        else:
            atualizar_log(f"Falha ao notificar Discord cobrança: HTTP {response.status_code} - {response.text}", cor="vermelho")
    except Exception as e:
        atualizar_log(f"Falha ao notificar Discord cobrança: {e}", cor="vermelho")

def notificar_discord_certificado(total_enviados):
    """Envia notificação ao Discord via webhook específico de certificado."""
    webhook_url = os.getenv("DISCORD_WEBHOOK_CERTIFICADO")
    if not webhook_url:
        atualizar_log("Webhook de certificado não configurado no .env.", cor="vermelho")
        return
    cargo = "<@&1299045050881151006>"
    payload = {
        "content": f"{cargo} **AutoMessenger WhatsApp - ComuniCertificado** - Mensagens de certificado enviadas! Total: {total_enviados} mensagem(ns) enviada(s)."
    }
    try:
        response = requests.post(webhook_url, json=payload)
        if response.status_code == 204:
            atualizar_log("Notificação de certificado enviada ao Discord.", cor="verde")
        elif response.status_code == 403:
            atualizar_log("Falha ao notificar Discord certificado: Webhook inválido ou deletado (403 Forbidden).", cor="vermelho")
        else:
            atualizar_log(f"Falha ao notificar Discord certificado: HTTP {response.status_code} - {response.text}", cor="vermelho")
    except Exception as e:
        atualizar_log(f"Falha ao notificar Discord certificado: {e}", cor="vermelho")

def finalizar_programa():
    messagebox.showinfo("Processo Finalizado", "Processamento concluído!")
    botao_fechar.configure(state="normal")
    botao_iniciar.configure(state="normal")
    botao_iniciar_chrome.configure(state="normal")  # Reativar o botão de Chrome
    botao_agendar.configure(state="normal")  # Reativar agendamento

def finalizar_programa_agendado():
    """Finaliza o programa após processamento agendado e fecha o Chrome"""
    global driver_agendamento
    messagebox.showinfo("Processo Finalizado", "Processamento agendado concluído!")
    botao_fechar.configure(state="normal")
    botao_iniciar.configure(state="normal")
    botao_iniciar_chrome.configure(state="normal")
    botao_agendar.configure(state="normal")

    # Fechar o Chrome do agendamento
    fechar_chrome_agendamento()

def processar_dados_agendado(excel, modelo, linha_inicial):
    """Processa os dados usando o driver já aberto pelo agendamento"""
    global driver_agendamento

    # Iniciar timer de processamento
    tempo_inicio = time.time()
    atualizar_log("Timer iniciado.", cor="azul")

    driver = driver_agendamento

    if not driver:
        atualizar_log("Driver não encontrado. Tentando abrir novo Chrome...", cor="vermelho")
        url = "https://web.whatsapp.com"
        driver = abrir_chrome_com_url(url)
        if not driver:
            atualizar_log("Não foi possível abrir o Chrome. Processamento abortado.", cor="vermelho")
            finalizar_programa_agendado()
            return

    # Verificar se o driver ainda está ativo
    try:
        driver.current_url
    except:
        atualizar_log("Sessão expirada. Tentando reconectar...", cor="vermelho")
        url = "https://web.whatsapp.com"
        driver = abrir_chrome_com_url(url)
        if not driver:
            atualizar_log("Não foi possível reconectar. Processamento abortado.", cor="vermelho")
            finalizar_programa_agendado()
            return
        driver_agendamento = driver

    time.sleep(5)
    dados = ler_dados_excel(excel, modelo, linha_inicial)
    if not dados:
        atualizar_log("Nenhum dado para processar.", cor="vermelho")
        finalizar_programa_agendado()
        return

    total_linhas = openpyxl.load_workbook(excel).active.max_row - linha_inicial + 1
    processamento_cancelado = False

    if modelo == "Cobranca":
        codigos, nomes, telefones, valores, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        ultima_ocorrencia_tel = {}
        for i, tel in enumerate(telefones):
            ultima_ocorrencia_tel[tel] = i
        estimar_tempo(total_contatos)
        for i, (cod, nome_emp, tel, p, v, carta) in enumerate(zip(codigos, nomes, telefones, valores, vencimentos, cartas)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando empresa {cod} - {nome_emp}: Tel: {tel}, Aviso nº: {carta}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, valores=p, vencimentos=v, carta=carta, nome_empresa=nome_emp)
            eh_ultimo = (ultima_ocorrencia_tel[tel] == i)
            if enviar_mensagem(driver, tel, mensagem, cod, nome_emp, enviar_aviso=eh_ultimo):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] Mensagem enviada para {tel}\n")
            aguardar_intervalo_envio()

    elif modelo == "ComuniCertificado":
        codigos, nomes, telefones, cnpjs, vencimentos, cartas = extrair_dados(dados, modelo)
        total_contatos = len(codigos)
        ultima_ocorrencia_tel = {}
        for i, tel in enumerate(telefones):
            ultima_ocorrencia_tel[tel] = i
        estimar_tempo(total_contatos)
        for i, (cod, nome_emp, tel, c, v, carta) in enumerate(zip(codigos, nomes, telefones, cnpjs, vencimentos, cartas)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            linha_atual = linha_inicial + i
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"Linha: {linha_atual}")
            atualizar_log(f"\nProcessando empresa {cod} - {nome_emp}: Tel: {tel}, Aviso nº: {carta}\n", cor="azul")
            mensagem = mensagem_padrao(modelo, vencimentos=v, carta=carta, cnpj=c, nome_empresa=nome_emp)
            eh_ultimo = (ultima_ocorrencia_tel[tel] == i)
            if enviar_mensagem(driver, tel, mensagem, cod, nome_emp, enviar_aviso=eh_ultimo):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] Mensagem enviada para {tel}\n")
            aguardar_intervalo_envio()

    elif modelo == "ONE":
        telefones_lista, empresas_lista, caminhos_lista = extrair_dados(dados, modelo)
        total_contatos = len(telefones_lista)
        estimar_tempo(total_contatos)
        linha_atual = linha_inicial
        for i, (tel, empresas, caminhos) in enumerate(zip(telefones_lista, empresas_lista, caminhos_lista)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando tel {tel}: {num_empresas} empresas\n", cor="azul")
            for cod, emp, _ in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")
            nomes_empresas = [emp for _, emp, _ in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas)
            identificador = ", ".join(nomes_empresas)
            if enviar_mensagem(driver, tel, mensagem, tel, identificador, modelo, caminhos):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] Mensagem enviada para {tel} com {num_empresas} arquivos\n")
            aguardar_intervalo_envio()
            linha_atual += num_empresas

    elif modelo == "ALL_info":
        telefones_lista, empresas_lista, extras = extrair_dados(dados, modelo)
        total_contatos = len(telefones_lista)
        estimar_tempo(total_contatos)
        linha_atual = linha_inicial
        for i, (tel, empresas, extra_info) in enumerate(zip(telefones_lista, empresas_lista, extras)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            competencia = extra_info.get('competencia', '')
            log_extra = f" - Competencia: {competencia}" if competencia else ""
            atualizar_log(f"\nProcessando tel {tel}: {num_empresas} empresas{log_extra}\n", cor="azul")
            for emp in empresas:
                cnpj_log = f" - CNPJ: {emp.get('cnpj', '')}" if emp.get('cnpj') else ""
                atualizar_log(f"Empresa: {emp['codigo']} - {emp['empresa']}{cnpj_log}")
            nomes_empresas = [emp['empresa'] for emp in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas, competencia=competencia, empresas_info=empresas)
            identificador = ", ".join(nomes_empresas)
            if enviar_mensagem(driver, tel, mensagem, tel, identificador, modelo):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now()}] Mensagem enviada para {tel} com {num_empresas} empresa(s){log_extra}\n")
            aguardar_intervalo_envio()
            linha_atual += num_empresas

    else:  # Modelo ALL
        telefones_lista, empresas_lista = extrair_dados(dados, modelo)
        total_contatos = len(telefones_lista)
        estimar_tempo(total_contatos)
        linha_atual = linha_inicial

        arquivo_anexo = None
        if anexo_habilitado and anexo_habilitado.get() and caminho_anexo and caminho_anexo.get():
            arquivo_anexo = caminho_anexo.get()
            if os.path.exists(arquivo_anexo):
                atualizar_log(f"Anexo configurado: {arquivo_anexo}", cor="azul")
            else:
                atualizar_log(f"Arquivo anexo não encontrado: {arquivo_anexo}", cor="vermelho")
                arquivo_anexo = None

        for i, (tel, empresas) in enumerate(zip(telefones_lista, empresas_lista)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                processamento_cancelado = True
                break
            num_empresas = len(empresas)
            linha_atual_final = linha_atual + num_empresas - 1
            porcentagem = ((i + 1) / total_contatos) * 100
            atualizar_progresso(porcentagem, f"{linha_atual_final}/{total_linhas + linha_inicial - 1}")
            atualizar_log(f"\nProcessando tel {tel}: {num_empresas} empresas\n", cor="azul")
            for cod, emp in empresas:
                atualizar_log(f"Empresa: {cod} - {emp}")
            nomes_empresas = [emp for _, emp in empresas]
            mensagem = mensagem_padrao(modelo, nome_empresa=nomes_empresas)
            identificador = ", ".join(nomes_empresas)
            caminhos_envio = [arquivo_anexo] if arquivo_anexo else None
            if enviar_mensagem(driver, tel, mensagem, tel, identificador, modelo, caminhos_envio):
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    anexo_info = " + anexo" if arquivo_anexo else ""
                    f.write(f"[{datetime.now()}] Mensagem enviada para {tel} com {num_empresas} empresa(s){anexo_info}\n")
            aguardar_intervalo_envio()
            linha_atual += num_empresas

    # Exibir tempo de processamento
    atualizar_log(f"Tempo total de processamento: {formatar_tempo(tempo_inicio)}", cor="verde")

    if not processamento_cancelado:
        atualizar_progresso(100, "Concluído")
        atualizar_log("Processamento agendado finalizado!", cor="verde")

    # Sempre finalizar e fechar o Chrome, mesmo se cancelado
    finalizar_programa_agendado()

def abrir_log():
    if log_file_path and os.path.exists(log_file_path):
        os.startfile(log_file_path)
    else:
        messagebox.showinfo("Log não disponível", "Não há log para esta sessão.")

def inicializar_arquivo_log(modelo):
    global log_file_path
    log_dir = os.path.join(os.path.dirname(__file__), 'AutoMessengerWhatsa_Logs')
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_path = os.path.join(log_dir, f"{modelo}_log_{timestamp}.txt")
    with open(log_file_path, 'w', encoding='utf-8') as f:
        f.write(f"=== Log AutoMessenger WhatsApp - {timestamp} ===\n\n")
    return log_file_path

def atualizar_log(mensagem, cor=None):
    log_text.configure(state="normal")
    timestamp = datetime.now().strftime("[%H:%M:%S] ")
    if cor == "vermelho":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "vermelho")
    elif cor == "verde":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "verde")
    elif cor == "azul":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "azul")
    else:
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "preto")
    log_text.configure(state="disabled")
    log_text.see("end")
    if log_file_path and os.path.exists(log_file_path):
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(f"{timestamp}{mensagem}\n")

def atualizar_progresso(valor, texto=""):
    progresso.set(valor / 100)
    progresso_texto.configure(text=texto)
    janela.update_idletasks()

def iniciar_chrome_automacao():
    # Verificar se há agendamento ativo
    if agendamento_ativo or keep_alive_ativo:
        messagebox.showwarning("Atenção", "Há um agendamento ativo. Cancele o agendamento antes de abrir o Chrome manualmente.")
        return

    atualizar_log("Iniciando configuração do Chrome de automação...", cor="azul")
    url = "https://web.whatsapp.com"
    driver = abrir_chrome_com_url(url)
    if driver:
        atualizar_log("Chrome aberto no WhatsApp Web. Escaneie o QR Code se necessário e inicie o processamento.", cor="azul")
    else:
        atualizar_log("Falha ao abrir o Chrome de automação.", cor="vermelho")

# Funções de Agendamento
def agendar_processamento():
    global agendamento_ativo, contagem_regressiva_ativa, data_hora_agendada, driver_agendamento

    # Validar campos antes de agendar
    excel = caminho_excel.get()
    modelo = modelo_selecionado.get()
    if not excel or not modelo:
        messagebox.showwarning("Atenção", "Selecione um modelo e um arquivo Excel antes de agendar.")
        return

    try:
        linha = int(entrada_linha_inicial.get())
        if linha < 2:
            raise ValueError("Linha inicial deve ser >= 2")
    except ValueError:
        messagebox.showwarning("Atenção", "Linha inicial deve ser um número inteiro >= 2.")
        return

    # Obter data e hora do agendamento
    try:
        data_str = entrada_data.get().strip()
        hora_str = entrada_hora.get().strip()

        # Validar formato
        if not data_str or not hora_str:
            messagebox.showwarning("Atenção", "Preencha a data e hora do agendamento.")
            return

        # Normalizar data: aceita 02012025 ou 02/01/2025
        data_str = data_str.replace("/", "").replace("-", "").replace(".", "")
        if len(data_str) == 8 and data_str.isdigit():
            data_str = f"{data_str[:2]}/{data_str[2:4]}/{data_str[4:]}"

        # Normalizar hora: aceita 0830 ou 08:30
        hora_str = hora_str.replace(":", "").replace(".", "").replace("-", "")
        if len(hora_str) == 4 and hora_str.isdigit():
            hora_str = f"{hora_str[:2]}:{hora_str[2:]}"

        # Converter para datetime
        data_hora_str = f"{data_str} {hora_str}"
        data_hora_agendada = datetime.strptime(data_hora_str, "%d/%m/%Y %H:%M")

        # Verificar se a data é futura
        agora = datetime.now()
        if data_hora_agendada <= agora:
            messagebox.showwarning("Atenção", "A data/hora deve ser no futuro.")
            return

        # Calcular diferença em segundos
        diferenca = (data_hora_agendada - agora).total_seconds()

        # Cancelar agendamento anterior se existir
        if agendamento_ativo:
            agendamento_ativo.cancel()
            parar_keep_alive()
            fechar_chrome_agendamento()

        # Abrir Chrome e iniciar keep-alive para manter sessão ativa
        atualizar_log("Abrindo Chrome para manter sessão ativa durante o agendamento...", cor="azul")
        driver_agendamento = abrir_chrome_agendamento()

        if not driver_agendamento:
            messagebox.showerror("Erro", "Não foi possível abrir o Chrome. Agendamento cancelado.")
            return

        # Aguardar um pouco para garantir que a página carregou
        time.sleep(5)

        # Iniciar keep-alive (refresh a cada 30 minutos)
        iniciar_keep_alive()

        # Criar novo timer
        agendamento_ativo = threading.Timer(diferenca, executar_agendamento)
        agendamento_ativo.start()

        # Iniciar contagem regressiva
        contagem_regressiva_ativa = True
        atualizar_contagem_regressiva()

        # Log do agendamento
        atualizar_log(f"=" * 50, cor="azul")
        atualizar_log(f"AGENDAMENTO CRIADO COM SUCESSO!", cor="verde")
        atualizar_log(f"Data/Hora programada: {data_hora_agendada.strftime('%d/%m/%Y às %H:%M')}", cor="azul")
        atualizar_log(f"Modelo: {modelo}", cor="azul")
        atualizar_log(f"Excel: {excel}", cor="azul")
        atualizar_log(f"Linha inicial: {linha}", cor="azul")
        atualizar_log(f"Tempo até execução: {formatar_tempo_restante(diferenca)}", cor="azul")
        atualizar_log(f"Keep-alive ativo: Refresh a cada 30 minutos", cor="azul")
        atualizar_log(f"=" * 50, cor="azul")

        # Desabilitar botões
        botao_agendar.configure(state="disabled")
        botao_cancelar_agendamento.configure(state="normal")
        botao_iniciar.configure(state="disabled")
        botao_iniciar_chrome.configure(state="disabled")

        messagebox.showinfo("Agendamento", f"Processamento agendado para:\n{data_hora_agendada.strftime('%d/%m/%Y às %H:%M')}\n\nO Chrome foi aberto e fará refresh automático a cada 30 minutos para manter a sessão ativa.\n\nPor favor, faça login se necessário.")

    except ValueError as e:
        messagebox.showerror("Erro", f"Formato de data/hora inválido.\nUse: DD/MM/AAAA e HH:MM\n\nErro: {str(e)}")

def executar_agendamento():
    global contagem_regressiva_ativa, agendamento_ativo
    contagem_regressiva_ativa = False
    agendamento_ativo = None

    # Parar o keep-alive antes de iniciar o processamento
    parar_keep_alive()

    # Atualizar log
    atualizar_log(f"=" * 50, cor="verde")
    atualizar_log(f"AGENDAMENTO EXECUTANDO!", cor="verde")
    atualizar_log(f"Horário: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}", cor="verde")
    atualizar_log(f"=" * 50, cor="verde")

    # Resetar botões (precisa ser feito na thread principal)
    janela.after(0, lambda: botao_agendar.configure(state="normal"))
    janela.after(0, lambda: botao_cancelar_agendamento.configure(state="disabled"))
    janela.after(0, lambda: label_contagem.configure(text=""))

    # Iniciar processamento usando o driver existente
    janela.after(0, iniciar_processamento_agendado)

def iniciar_processamento_agendado():
    """Inicia o processamento usando o driver já aberto pelo agendamento"""
    global cancelar, driver_agendamento
    cancelar = False
    excel = caminho_excel.get()
    modelo = modelo_selecionado.get()

    if not excel or not modelo:
        messagebox.showwarning("Atenção", "Selecione um modelo e um arquivo Excel.")
        return

    try:
        linha = int(entrada_linha_inicial.get())
        if linha < 2:
            raise ValueError("Linha inicial deve ser >= 2")
    except ValueError:
        messagebox.showwarning("Atenção", "Linha inicial deve ser um número inteiro >= 2.")
        return

    atualizar_log("Iniciando processamento agendado...", cor="azul")
    botao_iniciar.configure(state="disabled")
    botao_iniciar_chrome.configure(state="disabled")
    inicializar_arquivo_log(modelo)

    # Usar o driver existente do agendamento
    thread = threading.Thread(target=processar_dados_agendado, args=(excel, modelo, linha))
    thread.start()

def cancelar_agendamento():
    global agendamento_ativo, contagem_regressiva_ativa, data_hora_agendada

    if agendamento_ativo:
        agendamento_ativo.cancel()
        agendamento_ativo = None

    # Parar keep-alive e fechar Chrome
    parar_keep_alive()
    fechar_chrome_agendamento()

    contagem_regressiva_ativa = False
    data_hora_agendada = None

    # Resetar interface
    botao_agendar.configure(state="normal")
    botao_cancelar_agendamento.configure(state="disabled")
    botao_iniciar.configure(state="normal")
    botao_iniciar_chrome.configure(state="normal")
    label_contagem.configure(text="")

    atualizar_log("Agendamento cancelado pelo usuário.", cor="vermelho")
    messagebox.showinfo("Agendamento", "Agendamento cancelado com sucesso.")

def atualizar_contagem_regressiva():
    global contagem_regressiva_ativa

    if not contagem_regressiva_ativa or not data_hora_agendada:
        return

    agora = datetime.now()
    diferenca = (data_hora_agendada - agora).total_seconds()

    if diferenca <= 0:
        label_contagem.configure(text="Iniciando...")
        return

    # Formatar tempo restante
    texto = formatar_tempo_restante(diferenca)
    label_contagem.configure(text=f"Tempo restante: {texto}")

    # Atualizar a cada segundo
    janela.after(1000, atualizar_contagem_regressiva)

def formatar_tempo_restante(segundos):
    dias = int(segundos // 86400)
    horas = int((segundos % 86400) // 3600)
    minutos = int((segundos % 3600) // 60)
    segs = int(segundos % 60)

    partes = []
    if dias > 0:
        partes.append(f"{dias}d")
    if horas > 0:
        partes.append(f"{horas}h")
    if minutos > 0:
        partes.append(f"{minutos}m")
    partes.append(f"{segs}s")

    return " ".join(partes)

# Funções de Keep-Alive
def iniciar_keep_alive():
    """Inicia o sistema de keep-alive que faz refresh periódico no Chrome"""
    global keep_alive_ativo
    keep_alive_ativo = True
    atualizar_log("Keep-alive iniciado. Refresh a cada 30 minutos.", cor="azul")
    # Agendar primeiro refresh em 30 minutos (não fazer refresh imediato pois o Chrome acabou de abrir)
    janela.after(KEEP_ALIVE_INTERVALO, executar_keep_alive)

def executar_keep_alive():
    """Executa o refresh periódico para manter a sessão ativa"""
    global keep_alive_ativo, driver_agendamento

    if not keep_alive_ativo or not driver_agendamento:
        return

    def fazer_refresh():
        global keep_alive_ativo, driver_agendamento
        try:
            # Verificar se o driver ainda está ativo
            driver_agendamento.current_url

            # Fazer refresh na página
            driver_agendamento.refresh()
            atualizar_log(f"[Keep-alive] Refresh executado às {datetime.now().strftime('%H:%M:%S')}", cor="azul")

            # Agendar próximo refresh (30 minutos) - feito na thread principal
            if keep_alive_ativo:
                janela.after(KEEP_ALIVE_INTERVALO, executar_keep_alive)

        except Exception as e:
            atualizar_log(f"[Keep-alive] Erro no refresh: {str(e)}", cor="vermelho")
            # Tentar reconectar
            try:
                reconectar_chrome_agendamento()
                if keep_alive_ativo:
                    janela.after(KEEP_ALIVE_INTERVALO, executar_keep_alive)
            except:
                atualizar_log("[Keep-alive] Falha ao reconectar. Sessão pode ter expirado.", cor="vermelho")

    # Executar em thread separada para não travar a UI
    thread = threading.Thread(target=fazer_refresh, daemon=True)
    thread.start()

def parar_keep_alive():
    """Para o sistema de keep-alive"""
    global keep_alive_ativo
    keep_alive_ativo = False
    atualizar_log("Keep-alive parado.", cor="azul")

def abrir_chrome_agendamento():
    """Abre o Chrome para o agendamento e retorna o driver"""
    global driver_agendamento

    url = "https://web.whatsapp.com"

    # Encerra apenas o Chrome do perfil atual antes de abrir
    encerrar_processos_chrome()

    user_data_dir = obter_user_data_dir()
    perfil = obter_perfil_chrome()

    # Criar diretório se não existir
    if not os.path.exists(user_data_dir):
        os.makedirs(user_data_dir, exist_ok=True)
        atualizar_log(f"Diretório do perfil {perfil} criado.", cor="azul")

    atualizar_log(f"Abrindo Chrome para agendamento (Perfil: {perfil})...", cor="azul")

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-translate")
    chrome_options.add_argument("--lang=pt-BR")
    chrome_options.add_argument("--enable-javascript")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")

    service = Service(ChromeDriverManager().install())
    try:
        driver_agendamento = webdriver.Chrome(service=service, options=chrome_options)
        driver_agendamento.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
        driver_agendamento.set_page_load_timeout(180)
        driver_agendamento.get(url)
        atualizar_log(f"Chrome aberto no WhatsApp Web.", cor="verde")
        atualizar_log("Escaneie o QR Code se necessário.", cor="azul")

        return driver_agendamento
    except Exception as e:
        atualizar_log(f"Erro ao abrir Chrome: {str(e)}", cor="vermelho")
        driver_agendamento = None
        return None

def reconectar_chrome_agendamento():
    """Tenta reconectar o Chrome caso a sessão tenha caído"""
    global driver_agendamento

    atualizar_log("Tentando reconectar Chrome...", cor="azul")

    try:
        if driver_agendamento:
            driver_agendamento.quit()
    except:
        pass

    driver_agendamento = abrir_chrome_agendamento()
    if driver_agendamento:
        atualizar_log("Chrome reconectado com sucesso!", cor="verde")
    else:
        atualizar_log("Falha ao reconectar Chrome.", cor="vermelho")

def fechar_chrome_agendamento():
    """Fecha o Chrome do agendamento"""
    global driver_agendamento, keep_alive_ativo

    keep_alive_ativo = False

    if driver_agendamento:
        try:
            driver_agendamento.quit()
            atualizar_log("Chrome do agendamento fechado.", cor="azul")
        except:
            pass
        driver_agendamento = None

# Interface Principal
def main():
    global janela, caminho_excel, modelo_selecionado, mensagem_selecionada, botao_iniciar, botao_fechar, log_text, progresso, progresso_texto, entrada_linha_inicial, botao_iniciar_chrome, anexo_habilitado, caminho_anexo
    global entrada_data, entrada_hora, botao_agendar, botao_cancelar_agendamento, label_contagem
    global perfil_selecionado, botao_tema

    # Constantes de estilo compacto
    H_INPUT = 28
    H_BTN = 28
    H_BTN_ACTION = 30
    PAD_X = 12
    PAD_Y_ROW = 6
    FONT_LABEL = ("Segoe UI", 10)
    FONT_TITLE = ("Segoe UI", 11, "bold")
    FONT_HEADER = ("Segoe UI", 14, "bold")

    janela = ctk.CTk()
    janela.title("AutoMessenger WhatsApp")
    janela.geometry("880x380")
    janela.resizable(True, True)
    janela.minsize(800, 340)
    janela.protocol("WM_DELETE_WINDOW", fechar_programa)

    def resource_path(relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    try:
        janela.iconbitmap(resource_path(os.path.join("assets", "favicon.ico")))
    except:
        try:
            icon_image = ctk.CTkImage(Image.open(resource_path(os.path.join("assets", "favicon-32x32.png"))), size=(32, 32))
            janela.iconphoto(False, icon_image)
        except Exception as e:
            print(f"Falha ao carregar ícone: {e}")

    caminho_excel = ctk.StringVar()
    modelo_selecionado = ctk.StringVar()
    mensagem_selecionada = ctk.StringVar()
    progresso = ctk.DoubleVar()

    # ==================== LAYOUT 2 COLUNAS COMPACTO ====================

    # ========== HEADER COMPACTO ==========
    frame_header = ctk.CTkFrame(janela, fg_color="transparent", height=32)
    frame_header.pack(fill="x", padx=PAD_X, pady=(8, 6))
    frame_header.pack_propagate(False)

    try:
        logo_image = ctk.CTkImage(Image.open(resource_path(os.path.join("assets", "logo_ONE_Whatsa.png"))), size=(24, 24))
        logo_label = ctk.CTkLabel(frame_header, image=logo_image, text="")
        logo_label.pack(side="left", padx=(0, 8))
    except Exception as e:
        print(f"Error loading logo image: {e}")

    titulo = ctk.CTkLabel(frame_header, text="AutoMessenger WhatsApp", font=FONT_HEADER)
    titulo.pack(side="left")

    label_versao = ctk.CTkLabel(frame_header, text="v3.0 | Hugo L. Almeida", text_color="gray", font=("Segoe UI", 9))
    label_versao.pack(side="right", padx=(8, 0))

    # Botão de alternar tema (ao lado do autor)
    botao_tema = ctk.CTkButton(
        frame_header,
        text="☀",
        command=alternar_tema,
        width=28,
        height=24,
        font=("Segoe UI", 12),
        fg_color="transparent",
        hover_color=("gray80", "gray30"),
        text_color=("gray20", "gray80")
    )
    botao_tema.pack(side="right")

    # ========== CONTAINER PRINCIPAL (2 COLUNAS) ==========
    frame_principal = ctk.CTkFrame(janela, fg_color="transparent")
    frame_principal.pack(fill="both", expand=True, padx=PAD_X, pady=(0, 8))
    frame_principal.grid_columnconfigure(0, weight=55, uniform="col")
    frame_principal.grid_columnconfigure(1, weight=45, uniform="col")
    frame_principal.grid_rowconfigure(0, weight=1)

    # ==================== COLUNA ESQUERDA ====================
    frame_esquerda = ctk.CTkFrame(frame_principal, corner_radius=8)
    frame_esquerda.grid(row=0, column=0, sticky="nsew", padx=(0, 6), pady=0)

    # ----- Seção: Configurações -----
    label_config = ctk.CTkLabel(frame_esquerda, text="Configurações", font=FONT_TITLE)
    label_config.pack(anchor="w", padx=PAD_X, pady=(PAD_X, 8))

    # Linha 1: Modelo + Perfil + Chrome
    frame_row1 = ctk.CTkFrame(frame_esquerda, fg_color="transparent")
    frame_row1.pack(fill="x", padx=PAD_X, pady=(0, PAD_Y_ROW))

    ctk.CTkLabel(frame_row1, text="Modelo", font=FONT_LABEL, text_color="gray").pack(side="left")
    combo_modelo = ctk.CTkComboBox(frame_row1, values=list(MODELOS.keys()), variable=modelo_selecionado, width=110, height=H_INPUT, font=FONT_LABEL)
    combo_modelo.pack(side="left", padx=(6, 16))
    modelo_selecionado.trace_add("write", lambda *args: atualizar_mensagem_padrao())

    ctk.CTkLabel(frame_row1, text="Perfil", font=FONT_LABEL, text_color="gray").pack(side="left")
    perfil_selecionado = ctk.StringVar(value="1")
    combo_perfil = ctk.CTkComboBox(frame_row1, values=["1", "2", "Teste"], variable=perfil_selecionado, width=70, height=H_INPUT, font=FONT_LABEL)
    combo_perfil.pack(side="left", padx=(6, 16))

    botao_iniciar_chrome = ctk.CTkButton(frame_row1, text="Chrome Automação", command=iniciar_chrome_automacao, width=70, height=H_BTN, font=FONT_LABEL, fg_color="#4a5568", hover_color="#2d3748")
    botao_iniciar_chrome.pack(side="left")


    # Linha 2: Excel
    frame_row2 = ctk.CTkFrame(frame_esquerda, fg_color="transparent")
    frame_row2.pack(fill="x", padx=PAD_X, pady=(0, PAD_Y_ROW))

    ctk.CTkLabel(frame_row2, text="Excel", font=FONT_LABEL, text_color="gray", width=42).pack(side="left")
    entrada_excel = ctk.CTkEntry(frame_row2, textvariable=caminho_excel, height=H_INPUT, font=FONT_LABEL, placeholder_text="Selecione o arquivo...")
    entrada_excel.pack(side="left", padx=(6, 6), fill="x", expand=True)
    botao_excel = ctk.CTkButton(frame_row2, text="...", command=selecionar_excel, width=30, height=H_BTN, font=FONT_LABEL, fg_color="#4a5568", hover_color="#2d3748")
    botao_excel.pack(side="left")

    # Linha 3: Linha + Mensagem + Edit
    frame_row3 = ctk.CTkFrame(frame_esquerda, fg_color="transparent")
    frame_row3.pack(fill="x", padx=PAD_X, pady=(0, PAD_Y_ROW))

    ctk.CTkLabel(frame_row3, text="Linha", font=FONT_LABEL, text_color="gray").pack(side="left")
    entrada_linha_inicial = ctk.CTkEntry(frame_row3, width=40, height=H_INPUT, font=FONT_LABEL)
    entrada_linha_inicial.pack(side="left", padx=(6, 16))
    entrada_linha_inicial.insert(0, "2")

    ctk.CTkLabel(frame_row3, text="Msg", font=FONT_LABEL, text_color="gray").pack(side="left")
    mensagens = carregar_mensagens()
    combo_mensagem = ctk.CTkComboBox(frame_row3, values=list(mensagens.keys()), variable=mensagem_selecionada, width=130, height=H_INPUT, font=FONT_LABEL)
    combo_mensagem.pack(side="left", padx=(6, 6))
    mensagem_selecionada.set(list(mensagens.keys())[0])

    def abrir_editor_mensagem():
        janela_editor = ctk.CTkToplevel(janela)
        janela_editor.title("Editor de Mensagens")
        janela_editor.geometry("550x420")

        frame_editor = ctk.CTkFrame(janela_editor, fg_color="transparent")
        frame_editor.pack(fill="both", expand=True, padx=16, pady=16)

        # Dropdown para selecionar mensagem
        ctk.CTkLabel(frame_editor, text="Selecionar Mensagem:", font=FONT_LABEL).pack(anchor="w")
        mensagens_editor = carregar_mensagens()
        opcoes_dropdown = ["Nova Mensagem"] + list(mensagens_editor.keys())
        selecao_editor = ctk.StringVar(value="Nova Mensagem")
        dropdown_mensagens = ctk.CTkComboBox(frame_editor, values=opcoes_dropdown, variable=selecao_editor, width=300, height=H_INPUT, font=FONT_LABEL, state="readonly")
        dropdown_mensagens.pack(fill="x", pady=(2, 10))

        # Frame para nome (visível apenas para nova mensagem)
        frame_nome = ctk.CTkFrame(frame_editor, fg_color="transparent")
        frame_nome.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(frame_nome, text="Título:", font=FONT_LABEL).pack(anchor="w")
        entrada_nome = ctk.CTkEntry(frame_nome, height=H_INPUT, font=FONT_LABEL, placeholder_text="Digite o título da nova mensagem...")
        entrada_nome.pack(fill="x", pady=(2, 0))

        ctk.CTkLabel(frame_editor, text="Texto:", font=FONT_LABEL).pack(anchor="w")
        texto_mensagem = ctk.CTkTextbox(frame_editor, wrap="word", height=180, font=FONT_LABEL)
        texto_mensagem.pack(fill="both", expand=True, pady=(2, 12))

        def ao_selecionar_mensagem(escolha):
            texto_mensagem.delete("1.0", "end")
            entrada_nome.delete(0, "end")
            if escolha == "Nova Mensagem":
                frame_nome.pack(fill="x", pady=(0, 10), after=dropdown_mensagens)
                entrada_nome.configure(state="normal", placeholder_text="Digite o título da nova mensagem...")
            else:
                frame_nome.pack_forget()
                mensagens_atual = carregar_mensagens()
                if escolha in mensagens_atual:
                    texto_mensagem.insert("1.0", mensagens_atual[escolha])

        dropdown_mensagens.configure(command=ao_selecionar_mensagem)

        def salvar_mensagem():
            escolha = selecao_editor.get()
            texto = texto_mensagem.get("1.0", "end").strip()

            if escolha == "Nova Mensagem":
                nome = entrada_nome.get().strip()
                if not nome:
                    messagebox.showwarning("Atenção", "Digite o título da nova mensagem.")
                    return
            else:
                nome = escolha

            if not texto:
                messagebox.showwarning("Atenção", "O texto da mensagem é obrigatório.")
                return

            mensagens_atual = carregar_mensagens()
            if escolha == "Nova Mensagem" and nome in mensagens_atual:
                if not messagebox.askyesno("Confirmação", f"'{nome}' já existe. Sobrescrever?"):
                    return

            mensagens_atual[nome] = texto
            salvar_mensagens(mensagens_atual)
            combo_mensagem.configure(values=list(mensagens_atual.keys()))
            atualizar_log(f"Mensagem '{nome}' salva!", cor="verde")
            janela_editor.destroy()

        def remover_mensagem():
            escolha = selecao_editor.get()
            if escolha == "Nova Mensagem":
                messagebox.showwarning("Atenção", "Selecione uma mensagem existente para remover.")
                return

            mensagens_atual = carregar_mensagens()
            if escolha in mensagens_atual and messagebox.askyesno("Confirmação", f"Remover '{escolha}'?"):
                del mensagens_atual[escolha]
                salvar_mensagens(mensagens_atual)
                combo_mensagem.configure(values=list(mensagens_atual.keys()))
                mensagem_selecionada.set(list(mensagens_atual.keys())[0] if mensagens_atual else "")
                atualizar_log(f"Mensagem '{escolha}' removida!", cor="verde")
                janela_editor.destroy()

        frame_btns_editor = ctk.CTkFrame(frame_editor, fg_color="transparent")
        frame_btns_editor.pack(fill="x")
        ctk.CTkButton(frame_btns_editor, text="Salvar", command=salvar_mensagem, fg_color="#28a745", hover_color="#218838", width=90, height=H_BTN, font=FONT_LABEL).pack(side="left", padx=(0, 6))
        ctk.CTkButton(frame_btns_editor, text="Remover", command=remover_mensagem, fg_color="#dc3545", hover_color="#c82333", width=90, height=H_BTN, font=FONT_LABEL).pack(side="left")

    botao_editor = ctk.CTkButton(frame_row3, text="Editar/Remover", command=abrir_editor_mensagem, width=40, height=H_BTN, font=FONT_LABEL, fg_color="#4a5568", hover_color="#2d3748")
    botao_editor.pack(side="left")

    # Linha 4: Anexo
    frame_row4 = ctk.CTkFrame(frame_esquerda, fg_color="transparent")
    frame_row4.pack(fill="x", padx=PAD_X, pady=(0, 8))

    anexo_habilitado = ctk.BooleanVar(value=False)
    caminho_anexo = ctk.StringVar()

    checkbox_anexo = ctk.CTkCheckBox(frame_row4, text="Anexo", variable=anexo_habilitado, command=lambda: toggle_anexo(), width=20, height=H_INPUT, font=FONT_LABEL)
    checkbox_anexo.pack(side="left")
    entrada_anexo = ctk.CTkEntry(frame_row4, textvariable=caminho_anexo, state="disabled", height=H_INPUT, font=FONT_LABEL, placeholder_text="Arquivo...")
    entrada_anexo.pack(side="left", padx=(10, 6), fill="x", expand=True)

    def selecionar_anexo():
        arquivo = filedialog.askopenfilename(filetypes=[
            ("Vídeos", "*.mp4 *.avi *.mov *.mkv *.wmv *.webm"),
            ("Imagens", "*.jpg *.jpeg *.png *.gif *.bmp *.webp"),
            ("PDF", "*.pdf"),
            ("Documentos", "*.doc *.docx *.xls *.xlsx *.ppt *.pptx"),
            ("Todos", "*.*")
        ])
        if arquivo:
            caminho_anexo.set(arquivo)
            atualizar_log(f"Anexo: {arquivo}")

    botao_anexo = ctk.CTkButton(frame_row4, text="...", command=selecionar_anexo, state="disabled", width=30, height=H_BTN, font=FONT_LABEL, fg_color="#4a5568", hover_color="#2d3748")
    botao_anexo.pack(side="left")

    def toggle_anexo():
        if anexo_habilitado.get():
            entrada_anexo.configure(state="normal")
            botao_anexo.configure(state="normal")
        else:
            entrada_anexo.configure(state="disabled")
            botao_anexo.configure(state="disabled")
            caminho_anexo.set("")

    # ----- Separador fino -----
    ctk.CTkFrame(frame_esquerda, height=1, fg_color="#404040").pack(fill="x", padx=PAD_X, pady=(4, 10))

    # ----- Seção: Agendamento + Ações (unificado) -----
    frame_controles = ctk.CTkFrame(frame_esquerda, fg_color="transparent")
    frame_controles.pack(fill="x", padx=PAD_X, pady=(0, PAD_X))

    # Agendamento inline
    frame_agendar = ctk.CTkFrame(frame_controles, fg_color="transparent")
    frame_agendar.pack(fill="x", pady=(0, 8))

    ctk.CTkLabel(frame_agendar, text="Agendar", font=FONT_TITLE).pack(side="left", padx=(0, 10))
    entrada_data = ctk.CTkEntry(frame_agendar, width=85, height=H_INPUT, font=FONT_LABEL, placeholder_text="DD/MM/AAAA")
    entrada_data.pack(side="left", padx=(0, 6))
    entrada_hora = ctk.CTkEntry(frame_agendar, width=55, height=H_INPUT, font=FONT_LABEL, placeholder_text="HH:MM")
    entrada_hora.pack(side="left", padx=(0, 8))

    botao_agendar = ctk.CTkButton(frame_agendar, text="Agendar", command=agendar_processamento, fg_color="#6f42c1", hover_color="#5a32a3", width=70, height=H_BTN, font=FONT_LABEL)
    botao_agendar.pack(side="left", padx=(0, 4))
    botao_cancelar_agendamento = ctk.CTkButton(frame_agendar, text="Cancelar", command=cancelar_agendamento, fg_color="#fd7e14", hover_color="#e06b0a", width=70, height=H_BTN, font=FONT_LABEL, state="disabled")
    botao_cancelar_agendamento.pack(side="left")

    # Label de contagem regressiva (entre agendamento e ações)
    label_contagem = ctk.CTkLabel(frame_controles, text="", text_color="#6f42c1", font=("Segoe UI", 10, "bold"))
    label_contagem.pack(fill="x", pady=(4, 6))

    # Ações inline
    frame_acoes = ctk.CTkFrame(frame_controles, fg_color="transparent")
    frame_acoes.pack(fill="x")

    ctk.CTkLabel(frame_acoes, text="Ações", font=FONT_TITLE).pack(side="left", padx=(0, 14))
    botao_iniciar = ctk.CTkButton(frame_acoes, text="Iniciar", command=iniciar_processamento, fg_color="#28a745", hover_color="#218838", width=80, height=H_BTN_ACTION, font=FONT_LABEL)
    botao_iniciar.pack(side="left", padx=(0, 6))
    botao_cancelar = ctk.CTkButton(frame_acoes, text="Parar", command=cancelar_processamento, fg_color="#dc3545", hover_color="#c82333", width=70, height=H_BTN_ACTION, font=FONT_LABEL)
    botao_cancelar.pack(side="left", padx=(0, 6))
    botao_abrir_log = ctk.CTkButton(frame_acoes, text="Log", command=abrir_log, fg_color="#17a2b8", hover_color="#138496", width=60, height=H_BTN_ACTION, font=FONT_LABEL)
    botao_abrir_log.pack(side="left", padx=(0, 6))
    botao_fechar = ctk.CTkButton(frame_acoes, text="Fechar", command=fechar_programa, state="disabled", fg_color="#6c757d", hover_color="#5a6268", width=70, height=H_BTN_ACTION, font=FONT_LABEL)
    botao_fechar.pack(side="left")

    # ==================== COLUNA DIREITA: Log ====================
    frame_direita = ctk.CTkFrame(frame_principal, corner_radius=8)
    frame_direita.grid(row=0, column=1, sticky="nsew", padx=(6, 0), pady=0)

    # Header + Progresso inline
    frame_log_header = ctk.CTkFrame(frame_direita, fg_color="transparent")
    frame_log_header.pack(fill="x", padx=PAD_X, pady=(PAD_X, 8))

    label_log = ctk.CTkLabel(frame_log_header, text="Log", font=FONT_TITLE)
    label_log.pack(side="left", padx=(0, 12))

    barra_progresso = ctk.CTkProgressBar(frame_log_header, variable=progresso, height=6, width=120)
    barra_progresso.pack(side="left", padx=(0, 8))
    barra_progresso.set(0)

    progresso_texto = ctk.CTkLabel(frame_log_header, text="0/0", font=("Segoe UI", 9), text_color="gray")
    progresso_texto.pack(side="left")

    # Área de Log
    log_text = ctk.CTkTextbox(frame_direita, wrap="word", fg_color="#1a1a2e", corner_radius=6, font=("Consolas", 11))
    log_text.pack(fill="both", expand=True, padx=PAD_X, pady=(0, PAD_X))
    log_text.tag_config("vermelho", foreground="#ff6b6b")
    log_text.tag_config("verde", foreground="#51cf66")
    log_text.tag_config("azul", foreground="#74c0fc")
    log_text.tag_config("timestamp", foreground="#868e96")
    log_text.tag_config("preto", foreground="#ced4da")

    atualizar_log("AutoMessenger WhatsApp iniciado.", cor="verde")

    janela.mainloop()

if __name__ == '__main__':
    main()